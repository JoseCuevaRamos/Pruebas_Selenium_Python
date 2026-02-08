import os
from datetime import datetime
from typing import Optional

# Use openpyxl for Excel reporting
try:
    from openpyxl import Workbook, load_workbook
except Exception:
    Workbook = None
    load_workbook = None

DEFAULT_REPORT_NAME = "Test_Report.xlsx"


def get_report_path(root_dir: str, filename: str = DEFAULT_REPORT_NAME) -> str:
    return os.path.join(root_dir, filename)


def ensure_report_exists(path: str) -> None:
    """Create the Excel report with a 'Results' sheet and header if it doesn't exist yet."""
    if load_workbook is None:
        # openpyxl not installed; skip creation but allow callers to handle it
        return
    if not os.path.exists(path):
        wb = Workbook()
        ws = wb.active
        ws.title = "Results"
        ws.append(["Timestamp", "Suite", "CaseID", "Description", "Status", "Details"])
        wb.save(path)


def append_result(path: str, suite: str, case_id: str, description: str, status: str, details: Optional[str] = None) -> None:
    """Append a result row to the Excel report (creates file if missing)."""
    if load_workbook is None:
        # If openpyxl isn't available, silently ignore to avoid breaking tests
        return
    ensure_report_exists(path)
    wb = load_workbook(path)
    ws = wb["Results"] if "Results" in wb.sheetnames else wb.active
    ts = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    ws.append([ts, suite, case_id, description, status, details or ""])
    wb.save(path)
